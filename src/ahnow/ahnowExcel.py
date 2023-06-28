'''
Created on Jun 27, 2023

@author: billw
'''


from pathlib import Path
from openpyxl import load_workbook

class AHNowExcel(object):
    '''
    classdocs 
    
    Basic Excel operations on an existing Excel workbook
    '''

    def __init__(self, fullyQualifiedFileName, isReadOnly):
        '''
        Constructor
        '''
        self.fqfn = fullyQualifiedFileName 
        self.isReadOnly = isReadOnly
        self.wb = None
        self.ws = None
        
    def __del__(self):
        del self.wb
        del self.ws
        del self.isReadOnly
        del self.fqfn
        
    def isWorkbookPresent(self):
        filePath = Path(self.fqfn)
        return filePath.exists()
    
    def makeWorkbookFromTemplate(self, fqfnOfTemplate):        
        return None
        
    def getWorkbook(self):
        self.wb = load_workbook(self.fqfn, read_only=self.isReadOnly)
        return self.wb
        
    def getWorkSheet(self, worksheetName):
        self.ws = self.wb[worksheetName]
        return self.ws
    
    def listOfWorksheets(self):
        self.worksheetNames = self.wb.sheetnames
        return self.worksheetNames
    
    def getAllRows(self):
        rows = []
        for row in self.ws.rows:
            rows.append(row)
        return rows

    def saveAllRows(self):  
        self.wb.save(self.fqfn)
        
    def saveAs(self, fullyQualitifedFileName):
        self.wb.save(fullyQualitifedFileName)
                     
    def insertRows(self, idx, count):
        self.ws.insert_rows(idx, count)
    
    def addRows(self, rows):
        for row in rows:
            self.ws.append(row)
            
    def closeWorkbook(self):
        self.wb.close()
        
    def delAllEmptyRows(self, startPos, deleteCount):
        self.ws.delete_rows(startPos, deleteCount)
        

